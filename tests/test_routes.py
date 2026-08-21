"""Tests for Flask routes."""

import csv
import gzip
import importlib
import io
import json
import logging
import re
import tempfile
from unittest.mock import MagicMock, patch

import pytest
from flask import Response, request

import config as config_module
from app import check_rate_limit_topology, create_app, worker_count_from_start_command
from extensions import client_ip_key


class TestIndexRoute:
    def test_returns_200(self, client):
        response = client.get('/')
        assert response.status_code == 200

    def test_contains_html(self, client):
        response = client.get('/')
        assert b'JSON' in response.data


class TestHealthRoute:
    def test_returns_ok(self, client):
        response = client.get('/health')
        data = json.loads(response.data)
        assert data['status'] == 'ok'

    def test_returns_version(self, client, app):
        response = client.get('/health')
        data = json.loads(response.data)
        assert 'version' in data
        assert data['version'] == app.config['APP_VERSION']


class TestProcessRoute:
    def test_paste_valid_json(self, client):
        response = client.post(
            '/process',
            data={
                'input_method': 'paste',
                'pasted_json': '[{"id": 1, "name": "Alice"}, {"id": 2, "name": "Bob"}]',
                'json_path': '(root)',
            },
        )
        data = json.loads(response.data)
        assert data['success'] is True
        assert data['total_rows'] == 2
        assert 'id' in data['columns']
        assert 'name' in data['columns']

    def test_paste_invalid_json(self, client):
        response = client.post(
            '/process', data={'input_method': 'paste', 'pasted_json': '{invalid json}'}
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'error' in data

    def test_paste_empty(self, client):
        response = client.post('/process', data={'input_method': 'paste', 'pasted_json': ''})
        assert response.status_code == 400

    def test_invalid_input_method(self, client):
        response = client.post('/process', data={'input_method': 'unknown'})
        assert response.status_code == 400

    def test_nested_json_object(self, client):
        nested = json.dumps({'data': [{'x': 1}, {'x': 2}]})
        response = client.post(
            '/process', data={'input_method': 'paste', 'pasted_json': nested, 'json_path': 'data'}
        )
        data = json.loads(response.data)
        assert data['success'] is True
        assert data['total_rows'] == 2

    def test_no_path_returns_tree_payload(self, client):
        response = client.post(
            '/process', data={'input_method': 'paste', 'pasted_json': '[{"id": 1}, {"id": 2}]'}
        )
        data = json.loads(response.data)
        assert data.get('needs_selection') is True
        assert data['raw_json'] == [{'id': 1}, {'id': 2}]

    def test_file_upload(self, client):
        import io

        json_content = json.dumps([{'a': 1}])
        data = {
            'input_method': 'file',
            'json_path': '(root)',
            'json_file': (io.BytesIO(json_content.encode()), 'test.json'),
        }
        response = client.post('/process', data=data, content_type='multipart/form-data')
        result = json.loads(response.data)
        assert result['success'] is True

    def test_jsonl_paste(self, client):
        jsonl_content = '{"id": 1, "name": "Alice"}\n{"id": 2, "name": "Bob"}'
        response = client.post(
            '/process',
            data={
                'input_method': 'paste',
                'pasted_json': jsonl_content,
                'data_format': 'jsonl',
                'json_path': '(root)',
            },
        )
        data = json.loads(response.data)
        assert data['success'] is True
        assert data['total_rows'] == 2

    def test_jsonl_file_upload(self, client):
        import io

        jsonl_content = '{"a": 1}\n{"a": 2}\n{"a": 3}'
        data = {
            'input_method': 'file',
            'data_format': 'jsonl',
            'json_path': '(root)',
            'json_file': (io.BytesIO(jsonl_content.encode()), 'test.jsonl'),
        }
        response = client.post('/process', data=data, content_type='multipart/form-data')
        result = json.loads(response.data)
        assert result['success'] is True
        assert result['total_rows'] == 3

    def test_preview_limit(self, client, app):
        app.config['PREVIEW_ROW_LIMIT'] = 5
        rows = json.dumps([{'id': i} for i in range(20)])
        response = client.post(
            '/process', data={'input_method': 'paste', 'pasted_json': rows, 'json_path': '(root)'}
        )
        data = json.loads(response.data)
        assert data['total_rows'] == 20
        assert len(data['preview']) == 5


class TestExportCsvRoute:
    def test_export_valid_data(self, client):
        response = client.post(
            '/export-csv',
            data=json.dumps(
                {'csv_data': [{'a': 1, 'b': 2}, {'a': 3, 'b': 4}], 'csv_columns': ['a', 'b']}
            ),
            content_type='application/json',
        )
        assert response.status_code == 200
        assert response.content_type.startswith('text/csv')
        csv_text = response.data.decode('utf-8')
        assert 'a,b' in csv_text

    def test_export_empty_data(self, client):
        response = client.post(
            '/export-csv',
            data=json.dumps({'csv_data': [], 'csv_columns': []}),
            content_type='application/json',
        )
        assert response.status_code == 400


class TestExportXlsxRoute:
    def test_export_xlsx(self, client):
        response = client.post(
            '/export-xlsx',
            data=json.dumps({'csv_data': [{'a': 1, 'b': 2}], 'csv_columns': ['a', 'b']}),
            content_type='application/json',
        )
        assert response.status_code == 200
        assert 'spreadsheetml' in response.content_type

    def test_export_xlsx_empty(self, client):
        response = client.post(
            '/export-xlsx',
            data=json.dumps({'csv_data': [], 'csv_columns': []}),
            content_type='application/json',
        )
        assert response.status_code == 400


class TestPathSelection:
    def test_no_path_returns_raw_json_for_tree(self, client):
        payload = {'users': [{'n': 'A'}], 'orders': [{'id': 1}]}
        response = client.post(
            '/process', data={'input_method': 'paste', 'pasted_json': json.dumps(payload)}
        )
        data = json.loads(response.data)
        assert data.get('needs_selection') is True
        assert data['raw_json'] == payload

    def test_path_selection(self, client):
        multi = json.dumps({'users': [{'n': 'A'}], 'orders': [{'id': 1}, {'id': 2}]})
        response = client.post(
            '/process', data={'input_method': 'paste', 'pasted_json': multi, 'json_path': 'orders'}
        )
        data = json.loads(response.data)
        assert data['success'] is True
        assert data['total_rows'] == 2

    def test_path_to_array_index_object(self, client):
        payload = json.dumps({'data': [{'id': 1, 'orders': [{'x': 1}, {'x': 2}]}]})
        response = client.post(
            '/process',
            data={'input_method': 'paste', 'pasted_json': payload, 'json_path': 'data.0.orders'},
        )
        data = json.loads(response.data)
        assert data['success'] is True
        assert data['total_rows'] == 2

    def test_path_to_single_object_becomes_one_row(self, client):
        payload = json.dumps({'meta': {'version': 3, 'name': 'x'}})
        response = client.post(
            '/process', data={'input_method': 'paste', 'pasted_json': payload, 'json_path': 'meta'}
        )
        data = json.loads(response.data)
        assert data['success'] is True
        assert data['total_rows'] == 1
        assert 'version' in data['columns']

    def test_path_to_primitive_rejected(self, client):
        payload = json.dumps({'a': 1})
        response = client.post(
            '/process', data={'input_method': 'paste', 'pasted_json': payload, 'json_path': 'a'}
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'primitive' in data['error']

    def test_invalid_path_rejected(self, client):
        payload = json.dumps({'a': {'b': 1}})
        response = client.post(
            '/process', data={'input_method': 'paste', 'pasted_json': payload, 'json_path': 'a.c'}
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'not found' in data['error']


class TestApiFetch:
    def _mock_response(self, json_data, status_code=200):
        mock_resp = MagicMock()
        mock_resp.status_code = status_code
        content = json.dumps(json_data).encode()
        mock_resp.iter_content.return_value = [content]
        mock_resp.raise_for_status.return_value = None
        return mock_resp

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_api_fetch_success(self, mock_validate, mock_get, client):
        mock_validate.return_value = (True, None)
        mock_get.return_value = self._mock_response([{'id': 1}, {'id': 2}])

        response = client.post(
            '/process',
            data={
                'input_method': 'api',
                'api_url': 'https://api.example.com/data',
                'json_path': '(root)',
            },
        )
        data = json.loads(response.data)
        assert data['success'] is True
        assert data['total_rows'] == 2
        # Verify allow_redirects=False is passed
        _, kwargs = mock_get.call_args
        assert kwargs['allow_redirects'] is False

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_api_fetch_timeout(self, mock_validate, mock_get, client):
        import requests as req

        mock_validate.return_value = (True, None)
        mock_get.side_effect = req.exceptions.Timeout('timed out')

        response = client.post(
            '/process', data={'input_method': 'api', 'api_url': 'https://api.example.com/data'}
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'timed out' in data['error'].lower()

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_api_fetch_non_json(self, mock_validate, mock_get, client):
        mock_validate.return_value = (True, None)
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.iter_content.return_value = [b'<html>not json</html>']
        mock_resp.raise_for_status.return_value = None
        mock_get.return_value = mock_resp

        response = client.post(
            '/process', data={'input_method': 'api', 'api_url': 'https://api.example.com/data'}
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'not valid JSON' in data['error']

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_api_fetch_max_size_exceeded(self, mock_validate, mock_get, client, app):
        mock_validate.return_value = (True, None)
        app.config['API_FETCH_MAX_RESPONSE'] = 100  # 100 bytes
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.iter_content.return_value = [b'x' * 200]
        mock_resp.raise_for_status.return_value = None
        mock_get.return_value = mock_resp

        response = client.post(
            '/process', data={'input_method': 'api', 'api_url': 'https://api.example.com/data'}
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'exceeds maximum size' in data['error']

    @patch('routes.validate_url')
    def test_api_fetch_ssrf_blocked(self, mock_validate, client):
        mock_validate.return_value = (
            False,
            'URLs pointing to private or internal networks are not allowed',
        )

        response = client.post(
            '/process', data={'input_method': 'api', 'api_url': 'http://169.254.169.254/metadata'}
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'private' in data['error'].lower()

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_api_fetch_jsonl(self, mock_validate, mock_get, client):
        mock_validate.return_value = (True, None)
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        content = b'{"id": 1, "name": "Alice"}\n{"id": 2, "name": "Bob"}'
        mock_resp.iter_content.return_value = [content]
        mock_resp.raise_for_status.return_value = None
        mock_get.return_value = mock_resp

        response = client.post(
            '/process',
            data={
                'input_method': 'api',
                'api_url': 'https://api.example.com/data',
                'data_format': 'jsonl',
                'json_path': '(root)',
            },
        )
        data = json.loads(response.data)
        assert data['success'] is True
        assert data['total_rows'] == 2

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_api_fetch_request_error_no_leak(self, mock_validate, mock_get, client):
        import requests as req

        mock_validate.return_value = (True, None)
        mock_get.side_effect = req.exceptions.ConnectionError(
            'Connection to secret-internal-host:8080 refused'
        )

        response = client.post(
            '/process', data={'input_method': 'api', 'api_url': 'https://api.example.com/data'}
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        # Should NOT leak the internal connection details
        assert 'secret-internal-host' not in data['error']
        assert data['error'] == 'API request failed'


class TestFileUploadEncoding:
    def test_non_utf8_file_returns_400(self, client):
        import io

        # Latin-1 encoded content with bytes invalid in UTF-8
        content = b'\xff\xfe This is not valid UTF-8'
        data = {'input_method': 'file', 'json_file': (io.BytesIO(content), 'test.json')}
        response = client.post('/process', data=data, content_type='multipart/form-data')
        assert response.status_code == 400
        result = json.loads(response.data)
        assert 'UTF-8' in result['error']


class TestExportEdgeCases:
    def test_export_csv_no_json_body(self, client):
        response = client.post('/export-csv', data='not json', content_type='application/json')
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'Invalid or missing' in data['error']

    def test_export_xlsx_no_json_body(self, client):
        response = client.post('/export-xlsx', data='not json', content_type='application/json')
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'Invalid or missing' in data['error']

    def test_export_csv_no_content_type(self, client):
        response = client.post('/export-csv', data='hello')
        assert response.status_code == 400

    def test_export_xlsx_no_content_type(self, client):
        response = client.post('/export-xlsx', data='hello')
        assert response.status_code == 400


class TestSecurityHeaders:
    def test_csp_header(self, client):
        response = client.get('/')
        assert 'Content-Security-Policy' in response.headers

    def test_xframe_header(self, client):
        response = client.get('/')
        assert response.headers['X-Frame-Options'] == 'DENY'

    def test_xcontent_type_header(self, client):
        response = client.get('/')
        assert response.headers['X-Content-Type-Options'] == 'nosniff'

    def test_referrer_policy(self, client):
        response = client.get('/')
        assert 'strict-origin' in response.headers['Referrer-Policy']

    def test_permissions_policy(self, client):
        response = client.get('/')
        policy = response.headers['Permissions-Policy']
        assert 'camera=()' in policy
        assert 'microphone=()' in policy
        assert 'geolocation=()' in policy

    def test_cross_origin_headers(self, client):
        response = client.get('/')
        assert response.headers['Cross-Origin-Opener-Policy'] == 'same-origin'
        assert response.headers['Cross-Origin-Resource-Policy'] == 'same-origin'

    def test_csp_hardening_directives(self, client):
        directives = {
            part.strip() for part in client.get('/').headers['Content-Security-Policy'].split(';')
        }
        assert "object-src 'none'" in directives
        assert "base-uri 'self'" in directives
        assert "frame-ancestors 'none'" in directives
        assert "form-action 'self'" in directives
        assert 'upgrade-insecure-requests' in directives

    def test_csp_still_allows_google_fonts_and_data_images(self, client):
        csp = client.get('/').headers['Content-Security-Policy']
        assert 'https://fonts.googleapis.com' in csp
        assert 'https://fonts.gstatic.com' in csp
        assert 'data:' in csp

    def test_csp_has_no_malformed_directive(self, client):
        """A missing separator would fuse two directives into one token."""
        csp = client.get('/').headers['Content-Security-Policy']
        parts = [part.strip() for part in csp.split(';') if part.strip()]
        assert len(parts) == len(set(parts))
        for part in parts:
            assert not part.startswith("'")
            # Each directive starts with a bare directive name.
            assert re.match(r'^[a-z-]+( |$)', part), part

    def test_no_hsts_on_plain_http(self, client):
        response = client.get('/')
        assert 'Strict-Transport-Security' not in response.headers

    def test_hsts_on_secure_request(self, client):
        response = client.get('/', base_url='https://localhost')
        assert response.headers['Strict-Transport-Security'] == (
            'max-age=31536000; includeSubDomains'
        )


class TestFormulaInjection:
    """F1 - CSV/XLSX formula injection (CWE-1236)."""

    DANGEROUS = ['=SUM(A1)', '@cmd', '+1', '-1', '\tlead', '\rlead', '\nlead']

    def test_csv_prefixes_every_trigger(self, client):
        rows = [{'v': value} for value in self.DANGEROUS]
        response = client.post(
            '/export-csv',
            data=json.dumps({'csv_data': rows, 'csv_columns': ['v']}),
            content_type='application/json',
        )
        assert response.status_code == 200

        parsed = list(csv.reader(io.StringIO(response.data.decode('utf-8'))))
        emitted = [row[0] for row in parsed[1:]]
        assert emitted == ["'" + value for value in self.DANGEROUS]

    def test_csv_leaves_safe_values_alone(self, client):
        response = client.post(
            '/export-csv',
            data=json.dumps(
                {
                    'csv_data': [{'a': 'plain', 'b': 5, 'c': 'user@example.com'}],
                    'csv_columns': ['a', 'b', 'c'],
                }
            ),
            content_type='application/json',
        )
        parsed = list(csv.reader(io.StringIO(response.data.decode('utf-8'))))
        assert parsed[1] == ['plain', '5', 'user@example.com']

    def test_csv_sanitizes_column_headers(self, client):
        response = client.post(
            '/export-csv',
            data=json.dumps({'csv_data': [{'=EVIL()': 1}], 'csv_columns': ['=EVIL()']}),
            content_type='application/json',
        )
        parsed = list(csv.reader(io.StringIO(response.data.decode('utf-8'))))
        assert parsed[0] == ["'=EVIL()"]

    def test_csv_serializes_containers(self, client):
        response = client.post(
            '/export-csv',
            data=json.dumps({'csv_data': [{'a': {'k': 'v'}}], 'csv_columns': ['a']}),
            content_type='application/json',
        )
        parsed = list(csv.reader(io.StringIO(response.data.decode('utf-8'))))
        assert parsed[1] == ['{"k": "v"}']

    def test_xlsx_writes_triggers_as_string_cells(self, client):
        from openpyxl import load_workbook

        rows = [{'v': value} for value in self.DANGEROUS]
        response = client.post(
            '/export-xlsx',
            data=json.dumps({'csv_data': rows, 'csv_columns': ['v']}),
            content_type='application/json',
        )
        assert response.status_code == 200

        ws = load_workbook(io.BytesIO(response.data)).active
        for index, value in enumerate(self.DANGEROUS, start=2):
            cell = ws.cell(row=index, column=1)
            assert cell.data_type == 's', f'{value!r} was written as {cell.data_type}'
            # XLSX carries an explicit type, so the text itself stays intact.
            assert cell.value == value.replace('\r', '\n')

    def test_xlsx_sanitizes_column_headers(self, client):
        from openpyxl import load_workbook

        response = client.post(
            '/export-xlsx',
            data=json.dumps({'csv_data': [{'=EVIL()': 1}], 'csv_columns': ['=EVIL()']}),
            content_type='application/json',
        )
        ws = load_workbook(io.BytesIO(response.data)).active
        assert ws.cell(row=1, column=1).data_type == 's'

    def test_xlsx_keeps_numbers_numeric(self, client):
        from openpyxl import load_workbook

        response = client.post(
            '/export-xlsx',
            data=json.dumps({'csv_data': [{'a': -1, 'b': 2.5}], 'csv_columns': ['a', 'b']}),
            content_type='application/json',
        )
        ws = load_workbook(io.BytesIO(response.data)).active
        assert ws.cell(row=2, column=1).value == -1
        assert ws.cell(row=2, column=2).value == 2.5


class TestApiFetchLogHygiene:
    """F3/F9 - no URL component or token may reach the logs."""

    SECRET = 'sup3rs3cr3t-token'

    def _post(self, client, url, **extra):
        data = {'input_method': 'api', 'api_url': url}
        data.update(extra)
        return client.post('/process', data=data)

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_token_in_query_string_never_logged(self, mock_validate, mock_get, client, caplog):
        import requests as req

        mock_validate.return_value = (True, None)
        url = f'https://api.example.com/data?api_key={self.SECRET}'
        # requests puts the whole URL in the exception message.
        mock_get.side_effect = req.exceptions.ConnectionError(
            f"HTTPSConnectionPool(host='api.example.com', port=443): "
            f'Max retries exceeded with url: /data?api_key={self.SECRET}'
        )

        with caplog.at_level(logging.DEBUG):
            response = self._post(client, url)

        assert response.status_code == 400
        assert json.loads(response.data)['error'] == 'API request failed'

        logged = '\n'.join(record.getMessage() for record in caplog.records)
        assert self.SECRET not in logged
        assert 'api.example.com' not in logged
        assert '/data' not in logged
        assert 'API request failed' in logged

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_token_in_path_never_logged(self, mock_validate, mock_get, client, caplog):
        import requests as req

        mock_validate.return_value = (True, None)
        url = f'https://api.example.com/v1/{self.SECRET}/data'
        mock_get.side_effect = req.exceptions.ConnectionError(
            f'Failed to establish a new connection to /v1/{self.SECRET}/data'
        )

        with caplog.at_level(logging.DEBUG):
            response = self._post(client, url)

        assert response.status_code == 400
        logged = '\n'.join(record.getMessage() for record in caplog.records)
        assert self.SECRET not in logged
        assert 'v1' not in logged

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_query_param_auth_value_never_logged(self, mock_validate, mock_get, client, caplog):
        import requests as req

        mock_validate.return_value = (True, None)
        mock_get.side_effect = req.exceptions.ConnectionError('boom')

        with caplog.at_level(logging.DEBUG):
            response = self._post(
                client,
                'https://api.example.com/data',
                auth_method='query_param',
                query_param_name='api_key',
                query_param_value=self.SECRET,
            )

        assert response.status_code == 400
        assert self.SECRET not in '\n'.join(r.getMessage() for r in caplog.records)


class TestApiFetchJsonlErrors:
    """F9 - a malformed JSONL body from the API is a 400, not a logged 500."""

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_malformed_jsonl_returns_400(self, mock_validate, mock_get, client, caplog):
        mock_validate.return_value = (True, None)
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.iter_content.return_value = [b'{"a": 1}\n{bad json}\n']
        mock_resp.raise_for_status.return_value = None
        mock_get.return_value = mock_resp

        with caplog.at_level(logging.DEBUG):
            response = client.post(
                '/process',
                data={
                    'input_method': 'api',
                    'api_url': 'https://api.example.com/data',
                    'data_format': 'jsonl',
                },
            )

        assert response.status_code == 400
        assert json.loads(response.data)['error'] == 'API response is not valid JSONL'

        logged = '\n'.join(record.getMessage() for record in caplog.records)
        assert 'Unexpected error' not in logged
        assert 'bad json' not in logged
        assert not [r for r in caplog.records if r.levelno >= logging.ERROR]


class TestOutboundHeaderAllowlist:
    """F4 - the client supplies the outbound header NAME; only an allowlist passes."""

    REJECTED = [
        'Host',
        'host',
        'HOST',
        'Content-Length',
        'Transfer-Encoding',
        'Connection',
        'CoNnEcTiOn',
        'Proxy-Authorization',
        'PROXY-AUTHORIZATION',
        'Cookie',
        'X-CSRF-Token',
        ' Host ',
        'Host\t',
        'X Api Key',
        'X-Api-Key:',
        '',
    ]

    def _fetch(self, client, header_name):
        return client.post(
            '/process',
            data={
                'input_method': 'api',
                'api_url': 'https://api.example.com/data',
                'auth_method': 'api_key',
                'api_key_header': header_name,
                'api_key': 'secret',
                'json_path': '(root)',
            },
        )

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_reserved_and_hop_by_hop_names_rejected(self, mock_validate, mock_get, client):
        mock_validate.return_value = (True, None)

        for name in self.REJECTED:
            response = self._fetch(client, name)
            assert response.status_code == 400, f'{name!r} was accepted'
            assert 'not permitted' in json.loads(response.data)['error']

        # Nothing was ever sent upstream.
        assert mock_get.call_count == 0

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_permitted_header_in_unusual_case_is_forwarded(self, mock_validate, mock_get, client):
        mock_validate.return_value = (True, None)
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.iter_content.return_value = [b'[{"id": 1}]']
        mock_resp.raise_for_status.return_value = None
        mock_get.return_value = mock_resp

        response = self._fetch(client, '  x-API-kEy  ')
        assert json.loads(response.data)['success'] is True

        _, kwargs = mock_get.call_args
        assert kwargs['headers'] == {'x-API-kEy': 'secret'}

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_default_header_still_works(self, mock_validate, mock_get, client):
        mock_validate.return_value = (True, None)
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.iter_content.return_value = [b'[{"id": 1}]']
        mock_resp.raise_for_status.return_value = None
        mock_get.return_value = mock_resp

        response = self._fetch(client, 'X-API-Key')
        assert json.loads(response.data)['success'] is True
        _, kwargs = mock_get.call_args
        assert kwargs['headers'] == {'X-API-Key': 'secret'}


@pytest.fixture
def fresh_config(monkeypatch):
    """
    Rebuild config.Config from the current environment.

    Config holds class attributes evaluated at import time, so a monkeypatched
    env var only takes effect after a reload. F7 rules out constructing
    Config(...) -- it is a class, not a constructor.
    """

    def build(**env):
        for name, value in env.items():
            if value is None:
                monkeypatch.delenv(name, raising=False)
            else:
                monkeypatch.setenv(name, value)
        return importlib.reload(config_module).Config

    yield build
    # Leave the module holding the pristine values for every later test.
    monkeypatch.undo()
    importlib.reload(config_module)


class TestProductionSecretKey:
    """F7 - the dev SECRET_KEY must not survive into production."""

    def test_production_with_default_key_refuses_to_start(self, fresh_config):
        cfg = fresh_config(APP_ENV='production', SECRET_KEY=None)
        with pytest.raises(RuntimeError, match='SECRET_KEY must be set'):
            create_app(cfg)

    def test_production_with_empty_key_refuses_to_start(self, fresh_config):
        # A misconfigured secrets manager produces '' rather than "unset".
        cfg = fresh_config(APP_ENV='production', SECRET_KEY='')
        with pytest.raises(RuntimeError, match='SECRET_KEY must be set'):
            create_app(cfg)

    def test_production_with_real_key_starts(self, fresh_config):
        cfg = fresh_config(
            APP_ENV='production',
            SECRET_KEY='a-real-random-value',
            # A production app must also declare its topology (2.10).
            WEB_CONCURRENCY='1',
            APP_REPLICAS='1',
        )
        app = create_app(cfg)
        assert app.config['SECRET_KEY'] == 'a-real-random-value'

    def test_local_run_with_default_key_starts(self, fresh_config):
        # `python app.py` has DEBUG False, so gating on `not DEBUG` would have
        # blocked the documented local run.
        cfg = fresh_config(APP_ENV=None, SECRET_KEY=None, FLASK_DEBUG=None)
        assert cfg.DEBUG is False
        app = create_app(cfg)
        assert app.config['SECRET_KEY'] == config_module.DEV_SECRET_KEY

    def test_second_spelling_is_not_a_production_signal(self, fresh_config):
        # Accepting PRODUCTION=true as well would let a deployment pass this gate
        # while SESSION_COOKIE_SECURE (F16) stayed off.
        cfg = fresh_config(APP_ENV=None, PRODUCTION='true', SECRET_KEY=None)
        assert config_module.is_production() is False
        create_app(cfg)

    def test_app_env_is_case_and_space_insensitive(self, fresh_config):
        fresh_config(APP_ENV='  Production  ')
        assert config_module.is_production() is True


class TestIntegerConfigValidation:
    """F7 - a mistyped integer setting must name the variable, not raise ValueError."""

    INT_SETTINGS = [
        'MAX_UPLOAD_SIZE',
        'PREVIEW_ROW_LIMIT',
        'API_FETCH_TIMEOUT',
        'API_FETCH_MAX_RESPONSE',
        'FLATTEN_MAX_DEPTH',
        'API_DNS_TIMEOUT',
        'API_DNS_MAX_WORKERS',
        'API_DNS_ADMISSION_TIMEOUT',
    ]

    def test_each_integer_setting_reports_a_clear_error(self, fresh_config):
        for name in self.INT_SETTINGS:
            with pytest.raises(RuntimeError, match=f'{name} must be an integer'):
                fresh_config(**{name: 'abc'})
            # Undo before the next iteration so errors do not stack.
            fresh_config(**{name: None})

    def test_blank_value_falls_back_to_the_default(self, fresh_config):
        cfg = fresh_config(PREVIEW_ROW_LIMIT='   ')
        assert cfg.PREVIEW_ROW_LIMIT == 25

    def test_valid_value_is_applied(self, fresh_config):
        cfg = fresh_config(PREVIEW_ROW_LIMIT=' 7 ')
        assert cfg.PREVIEW_ROW_LIMIT == 7

    def test_port_allowlist_reports_a_clear_error(self, fresh_config):
        with pytest.raises(RuntimeError, match='API_ALLOWED_PORTS must be a comma-separated'):
            fresh_config(API_ALLOWED_PORTS='80,https')
        fresh_config(API_ALLOWED_PORTS=None)

    def test_port_allowlist_is_parsed(self, fresh_config):
        cfg = fresh_config(API_ALLOWED_PORTS=' 80 , 8443 ')
        assert sorted(cfg.API_ALLOWED_PORTS) == [80, 8443]


class TestRecursionDepth:
    """F8 - a pathologically nested document is a 400, never a 500."""

    @staticmethod
    def _deep_json(depth):
        return '{"a":' * depth + '1' + '}' * depth

    def test_deeply_nested_paste_returns_400(self, client, caplog):
        with caplog.at_level(logging.DEBUG):
            response = client.post(
                '/process',
                data={'input_method': 'paste', 'pasted_json': self._deep_json(1500)},
            )
        assert response.status_code == 400
        assert json.loads(response.data)['error'] == 'JSON nesting too deep'
        assert not [r for r in caplog.records if r.levelno >= logging.ERROR]

    def test_deeply_nested_upload_returns_400(self, client):
        response = client.post(
            '/process',
            data={
                'input_method': 'file',
                'json_file': (io.BytesIO(self._deep_json(1500).encode()), 'deep.json'),
            },
            content_type='multipart/form-data',
        )
        assert response.status_code == 400
        assert json.loads(response.data)['error'] == 'JSON nesting too deep'

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_deeply_nested_api_response_returns_400(self, mock_validate, mock_get, client):
        mock_validate.return_value = (True, None)
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.iter_content.return_value = [self._deep_json(1500).encode()]
        mock_resp.raise_for_status.return_value = None
        mock_get.return_value = mock_resp

        response = client.post(
            '/process',
            data={'input_method': 'api', 'api_url': 'https://api.example.com/data'},
        )
        assert response.status_code == 400

    def test_moderately_nested_document_still_works(self, client):
        payload = json.dumps({'rows': [{'id': 1}, {'id': 2}]})
        response = client.post(
            '/process',
            data={'input_method': 'paste', 'pasted_json': payload, 'json_path': 'rows'},
        )
        assert json.loads(response.data)['total_rows'] == 2


class TestProxyAwareRateLimiting:
    """F12 / D3 - X-Forwarded-For is honored only under TRUST_PROXY=1."""

    @staticmethod
    def _app_capturing_remote_addr(cfg):
        app = create_app(cfg)
        app.config['TESTING'] = True
        app.config['WTF_CSRF_ENABLED'] = False
        seen = {}

        @app.before_request
        def _capture():
            seen['remote_addr'] = request.remote_addr
            seen['key'] = client_ip_key()
            seen['is_secure'] = request.is_secure

        return app, seen

    def test_forwarded_for_ignored_by_default(self, fresh_config):
        cfg = fresh_config(TRUST_PROXY=None)
        app, seen = self._app_capturing_remote_addr(cfg)

        app.test_client().get(
            '/health',
            headers={'X-Forwarded-For': '9.9.9.9', 'X-Forwarded-Proto': 'https'},
            environ_base={'REMOTE_ADDR': '10.0.0.5'},
        )

        assert seen['remote_addr'] == '10.0.0.5'
        assert seen['key'] == '10.0.0.5'
        assert seen['is_secure'] is False

    def test_forwarded_for_used_when_trusted(self, fresh_config):
        cfg = fresh_config(TRUST_PROXY='1')
        app, seen = self._app_capturing_remote_addr(cfg)

        app.test_client().get(
            '/health',
            headers={'X-Forwarded-For': '9.9.9.9', 'X-Forwarded-Proto': 'https'},
            environ_base={'REMOTE_ADDR': '10.0.0.5'},
        )

        assert seen['remote_addr'] == '9.9.9.9'
        assert seen['key'] == '9.9.9.9'
        # x_proto=1 also makes is_secure correct behind a TLS-terminating proxy,
        # which HSTS (1.5) and the Secure cookie (1.14) depend on.
        assert seen['is_secure'] is True

    def test_only_one_hop_is_trusted(self, fresh_config):
        """
        A client that prepends its own hop must not choose its bucket.

        With x_for=1 ProxyFix takes the LAST entry -- the hop our single trusted
        proxy actually appended -- so the forged leading entry is ignored.
        """
        cfg = fresh_config(TRUST_PROXY='1')
        app, seen = self._app_capturing_remote_addr(cfg)

        app.test_client().get(
            '/health',
            headers={'X-Forwarded-For': '1.1.1.1, 2.2.2.2, 3.3.3.3'},
            environ_base={'REMOTE_ADDR': '10.0.0.5'},
        )

        assert seen['remote_addr'] == '3.3.3.3'

    def test_different_clients_get_different_buckets(self, fresh_config):
        cfg = fresh_config(TRUST_PROXY='1')
        app, seen = self._app_capturing_remote_addr(cfg)
        client = app.test_client()

        keys = []
        for ip in ('9.9.9.9', '8.8.8.8'):
            client.get(
                '/health',
                headers={'X-Forwarded-For': ip},
                environ_base={'REMOTE_ADDR': '10.0.0.5'},
            )
            keys.append(seen['key'])

        assert keys == ['9.9.9.9', '8.8.8.8']


class TestJsonErrorHandlers:
    """F10 - every error response is JSON, including the framework's own."""

    def test_oversized_request_returns_json_413(self, fresh_config):
        cfg = fresh_config(MAX_UPLOAD_SIZE=str(1024 * 1024))
        app = create_app(cfg)
        app.config['WTF_CSRF_ENABLED'] = False

        response = app.test_client().post(
            '/process',
            data={'input_method': 'paste', 'pasted_json': 'x' * (2 * 1024 * 1024)},
        )

        assert response.status_code == 413
        assert response.content_type.startswith('application/json')
        assert json.loads(response.data)['error'] == 'Request too large (max 1MB)'

    def test_unknown_route_returns_json_404(self, client):
        response = client.get('/no-such-route')
        assert response.status_code == 404
        assert response.content_type.startswith('application/json')
        assert 'error' in json.loads(response.data)

    def test_internal_error_returns_json_500(self, fresh_config):
        cfg = fresh_config()
        app = create_app(cfg)
        app.config['WTF_CSRF_ENABLED'] = False
        # TESTING would re-raise instead of routing to the handler.
        app.config['PROPAGATE_EXCEPTIONS'] = False

        @app.route('/boom')
        def _boom():
            raise RuntimeError('kaboom')

        response = app.test_client().get('/boom')
        assert response.status_code == 500
        assert response.content_type.startswith('application/json')
        assert json.loads(response.data)['error'] == 'An internal error occurred'
        assert b'kaboom' not in response.data


class TestNoStoreCacheControl:
    """F11 - data-bearing responses must not be retained by any cache."""

    def test_health_is_no_store(self, client):
        assert client.get('/health').headers['Cache-Control'] == 'no-store'

    def test_process_is_no_store(self, client):
        response = client.post(
            '/process',
            data={
                'input_method': 'paste',
                'pasted_json': '[{"a": 1}]',
                'json_path': '(root)',
            },
        )
        assert response.headers['Cache-Control'] == 'no-store'

    def test_export_csv_is_no_store(self, client):
        response = client.post(
            '/export-csv',
            data=json.dumps({'csv_data': [{'a': 1}], 'csv_columns': ['a']}),
            content_type='application/json',
        )
        assert response.headers['Cache-Control'] == 'no-store'

    def test_export_xlsx_is_no_store(self, client):
        response = client.post(
            '/export-xlsx',
            data=json.dumps({'csv_data': [{'a': 1}], 'csv_columns': ['a']}),
            content_type='application/json',
        )
        assert response.headers['Cache-Control'] == 'no-store'

    def test_index_page_is_still_cacheable(self, client):
        assert client.get('/').headers.get('Cache-Control') != 'no-store'


class TestUploadValidation:
    """F13 - the server, not just the file input's accept attribute."""

    def _upload(self, client, filename, content_type=None, body=b'[{"a": 1}]'):
        data = {
            'input_method': 'file',
            'json_path': '(root)',
            'json_file': (io.BytesIO(body), filename, content_type)
            if content_type is not None
            else (io.BytesIO(body), filename),
        }
        return client.post('/process', data=data, content_type='multipart/form-data')

    def test_rejects_unexpected_extensions(self, client):
        for filename in ('evil.txt', 'evil.exe', 'evil', 'evil.json.png', 'evil.csv'):
            response = self._upload(client, filename)
            assert response.status_code == 400, filename
            assert '.json or .jsonl' in json.loads(response.data)['error']

    def test_rejects_unexpected_content_type(self, client):
        response = self._upload(client, 'data.json', content_type='image/png')
        assert response.status_code == 400
        assert 'Unsupported content type' in json.loads(response.data)['error']

    def test_accepts_json_and_jsonl(self, client):
        assert json.loads(self._upload(client, 'data.json').data)['success'] is True
        assert json.loads(self._upload(client, 'DATA.JSON').data)['success'] is True

    def test_accepts_the_octet_stream_browsers_send_for_jsonl(self, client):
        response = self._upload(
            client,
            'data.jsonl',
            content_type='application/octet-stream',
            body=b'{"a": 1}',
        )
        assert response.status_code == 200


class TestCookieHardening:
    """F16 - explicit cookie flags, Secure tied to APP_ENV=production."""

    @staticmethod
    def _set_cookie(app):
        app.config['TESTING'] = True
        # The index page calls csrf_token(), which writes to the session.
        response = app.test_client().get('/', base_url='https://localhost')
        return response.headers.get('Set-Cookie', '')

    def test_local_run_flags(self, fresh_config):
        cfg = fresh_config(APP_ENV=None)
        assert cfg.SESSION_COOKIE_HTTPONLY is True
        assert cfg.SESSION_COOKIE_SAMESITE == 'Lax'
        assert cfg.SESSION_COOKIE_SECURE is False

        cookie = self._set_cookie(create_app(cfg))
        assert 'HttpOnly' in cookie
        assert 'SameSite=Lax' in cookie
        assert 'Secure' not in cookie

    def test_production_sets_secure(self, fresh_config):
        cfg = fresh_config(
            APP_ENV='production',
            SECRET_KEY='a-real-random-value',
            WEB_CONCURRENCY='1',
            APP_REPLICAS='1',
        )
        assert cfg.SESSION_COOKIE_SECURE is True

        cookie = self._set_cookie(create_app(cfg))
        assert 'Secure' in cookie
        assert 'HttpOnly' in cookie
        assert 'SameSite=Lax' in cookie


class TestHealthVersionGate:
    """F15 - version is returned by default; the gate only lets operators opt out."""

    def test_version_present_by_default(self, fresh_config):
        cfg = fresh_config(HEALTH_REVEAL_VERSION=None)
        assert cfg.HEALTH_REVEAL_VERSION is True

        app = create_app(cfg)
        data = json.loads(app.test_client().get('/health').data)
        assert data['status'] == 'ok'
        assert data['version'] == cfg.APP_VERSION

    def test_version_hidden_when_disabled(self, fresh_config):
        cfg = fresh_config(HEALTH_REVEAL_VERSION='0')
        assert cfg.HEALTH_REVEAL_VERSION is False

        app = create_app(cfg)
        data = json.loads(app.test_client().get('/health').data)
        assert data == {'status': 'ok'}


class TestGzipCompression:
    """P1 - compress large text/JSON bodies, and nothing else."""

    @staticmethod
    def _big_payload(rows=400):
        return json.dumps([{'id': i, 'name': f'user-{i}', 'note': 'x' * 60} for i in range(rows)])

    def _process(self, client, **headers):
        return client.post(
            '/process',
            data={
                'input_method': 'paste',
                'pasted_json': self._big_payload(),
                'json_path': '(root)',
            },
            headers=headers,
        )

    def test_large_json_is_compressed(self, client):
        response = self._process(client, **{'Accept-Encoding': 'gzip, deflate'})

        assert response.headers['Content-Encoding'] == 'gzip'
        assert 'Accept-Encoding' in response.headers['Vary']

        raw = response.get_data()
        decompressed = gzip.decompress(raw)
        assert json.loads(decompressed)['total_rows'] == 400
        assert len(raw) < len(decompressed)
        # Content-Length must describe what is actually on the wire.
        assert int(response.headers['Content-Length']) == len(raw)

    def test_not_compressed_without_accept_encoding(self, client):
        response = self._process(client, **{'Accept-Encoding': 'identity'})
        assert 'Content-Encoding' not in response.headers
        assert 'Accept-Encoding' in response.headers['Vary']
        assert json.loads(response.data)['total_rows'] == 400

    def test_small_response_is_not_compressed(self, client):
        response = client.get('/health', headers={'Accept-Encoding': 'gzip'})
        assert 'Content-Encoding' not in response.headers
        assert json.loads(response.data)['status'] == 'ok'

    def test_vary_is_not_duplicated(self, client):
        response = self._process(client, **{'Accept-Encoding': 'gzip'})
        varies = [v.strip().lower() for v in response.headers.get_all('Vary')]
        assert varies.count('accept-encoding') == 1

    def test_head_request_is_left_alone(self, client):
        response = client.head('/', headers={'Accept-Encoding': 'gzip'})
        assert 'Content-Encoding' not in response.headers

    def test_already_encoded_body_is_left_alone(self, app):
        app.config['PROPAGATE_EXCEPTIONS'] = False

        @app.route('/pre-encoded')
        def _pre_encoded():
            payload = gzip.compress(b'{"already": "' + b'x' * 5000 + b'"}')
            return Response(
                payload,
                mimetype='application/json',
                headers={'Content-Encoding': 'gzip'},
            )

        response = app.test_client().get('/pre-encoded', headers={'Accept-Encoding': 'gzip'})
        # Not double-compressed: one gzip layer decodes to the original JSON.
        assert response.headers['Content-Encoding'] == 'gzip'
        assert gzip.decompress(response.get_data()).startswith(b'{"already"')

    def test_streamed_response_is_left_alone(self, app):
        @app.route('/streamed')
        def _streamed():
            def generate():
                for index in range(500):
                    yield f'line {index} ' + 'y' * 40 + '\n'

            return Response(generate(), mimetype='text/plain')

        response = app.test_client().get('/streamed', headers={'Accept-Encoding': 'gzip'})
        assert 'Content-Encoding' not in response.headers
        assert response.get_data().startswith(b'line 0 ')

    def test_binary_export_is_not_compressed(self, client):
        response = client.post(
            '/export-xlsx',
            data=json.dumps({'csv_data': [{'a': 'x' * 100}] * 50, 'csv_columns': ['a']}),
            content_type='application/json',
            headers={'Accept-Encoding': 'gzip'},
        )
        # XLSX is a zip container; re-compressing it wastes CPU for nothing.
        assert 'Content-Encoding' not in response.headers


class TestStaticAssetCaching:
    """P6 - static assets are cacheable, and their URLs are version-busted."""

    def test_static_assets_carry_a_max_age(self, client):
        response = client.get('/static/css/style.css')
        assert response.status_code == 200
        assert 'max-age=86400' in response.headers['Cache-Control']

    def test_asset_urls_are_version_busted(self, client, app):
        html = client.get('/').data.decode('utf-8')
        version = app.config['APP_VERSION']
        assert f'css/style.css?v={version}' in html
        assert f'js/app.js?v={version}' in html

    def test_max_age_is_configurable(self, fresh_config):
        cfg = fresh_config(STATIC_MAX_AGE='60')
        app = create_app(cfg)
        response = app.test_client().get('/static/js/app.js')
        assert 'max-age=60' in response.headers['Cache-Control']


class TestStreamingCsvExport:
    """P3 - CSV is generator-streamed and stays uncapped."""

    def test_response_is_streamed(self, app):
        client = app.test_client()
        with app.test_request_context():
            pass
        response = client.post(
            '/export-csv',
            data=json.dumps(
                {
                    'csv_data': [{'a': i, 'b': 'x' * 20} for i in range(2000)],
                    'csv_columns': ['a', 'b'],
                }
            ),
            content_type='application/json',
        )
        assert response.status_code == 200
        # No Content-Length: the body is produced as it is written.
        assert 'Content-Length' not in response.headers
        rows = list(csv.reader(io.StringIO(response.get_data(as_text=True))))
        assert rows[0] == ['a', 'b']
        assert len(rows) == 2001

    def test_csv_is_not_capped_by_the_xlsx_budget(self, app):
        app.config['MAX_EXPORT_CELLS'] = 10
        response = app.test_client().post(
            '/export-csv',
            data=json.dumps({'csv_data': [{'a': i} for i in range(200)], 'csv_columns': ['a']}),
            content_type='application/json',
        )
        assert response.status_code == 200
        assert len(response.get_data(as_text=True).strip().splitlines()) == 201

    def test_chunk_boundary_does_not_lose_or_duplicate_rows(self, app):
        # Exercise exactly the flush boundary of CSV_STREAM_CHUNK_ROWS.
        from routes import CSV_STREAM_CHUNK_ROWS

        for count in (
            CSV_STREAM_CHUNK_ROWS - 1,
            CSV_STREAM_CHUNK_ROWS,
            CSV_STREAM_CHUNK_ROWS + 1,
        ):
            response = app.test_client().post(
                '/export-csv',
                data=json.dumps(
                    {'csv_data': [{'a': i} for i in range(count)], 'csv_columns': ['a']}
                ),
                content_type='application/json',
            )
            rows = list(csv.reader(io.StringIO(response.get_data(as_text=True))))
            assert [r[0] for r in rows[1:]] == [str(i) for i in range(count)], count


class TestXlsxExportBudget:
    """P3 / D6 - the XLSX guard is on by default, budgeted in cells, never silent."""

    def test_process_advertises_the_budget(self, client, app):
        app.config['MAX_EXPORT_CELLS'] = 1000
        response = client.post(
            '/process',
            data={
                'input_method': 'paste',
                'pasted_json': json.dumps([{'a': i, 'b': i} for i in range(5)]),
                'json_path': '(root)',
            },
        )
        data = json.loads(response.data)
        assert data['total_rows'] == 5
        assert data['total_cells'] == 10
        assert data['max_export_cells'] == 1000

    def test_existing_keys_are_unchanged(self, client):
        response = client.post(
            '/process',
            data={
                'input_method': 'paste',
                'pasted_json': '[{"a": 1}]',
                'json_path': '(root)',
            },
        )
        data = json.loads(response.data)
        for key in ('success', 'columns', 'preview', 'total_rows', 'csv_data', 'csv_columns'):
            assert key in data
        assert data['total_rows'] == 1
        assert data['columns'] == ['a']

    def test_oversized_export_is_refused_not_truncated(self, app):
        app.config['MAX_EXPORT_CELLS'] = 10
        response = app.test_client().post(
            '/export-xlsx',
            data=json.dumps(
                {'csv_data': [{'a': i, 'b': i} for i in range(20)], 'csv_columns': ['a', 'b']}
            ),
            content_type='application/json',
        )
        assert response.status_code == 400
        error = json.loads(response.data)['error']
        assert '40 cells' in error
        assert 'limit of 10' in error
        assert 'CSV or TSV' in error

    def test_export_at_the_limit_succeeds(self, app):
        app.config['MAX_EXPORT_CELLS'] = 40
        response = app.test_client().post(
            '/export-xlsx',
            data=json.dumps(
                {'csv_data': [{'a': i, 'b': i} for i in range(20)], 'csv_columns': ['a', 'b']}
            ),
            content_type='application/json',
        )
        assert response.status_code == 200

    def test_zero_disables_the_guard(self, app):
        app.config['MAX_EXPORT_CELLS'] = 0
        response = app.test_client().post(
            '/export-xlsx',
            data=json.dumps({'csv_data': [{'a': i} for i in range(50)], 'csv_columns': ['a']}),
            content_type='application/json',
        )
        assert response.status_code == 200

    def test_guard_is_enabled_by_default(self, app):
        assert app.config['MAX_EXPORT_CELLS'] > 0

    def test_export_writes_no_files(self, app, tmp_path, monkeypatch):
        """
        D6 / AGENTS.md: no disk writes of payloads.

        openpyxl's write_only mode and a rolled-over SpooledTemporaryFile both
        put payload bytes in the OS temp directory. Point every temp mechanism at
        an empty directory and assert nothing lands there.
        """
        for var in ('TMPDIR', 'TEMP', 'TMP'):
            monkeypatch.setenv(var, str(tmp_path))
        monkeypatch.setattr(tempfile, 'tempdir', str(tmp_path))

        response = app.test_client().post(
            '/export-xlsx',
            data=json.dumps(
                {
                    'csv_data': [{'a': f'value-{i}', 'b': i} for i in range(3000)],
                    'csv_columns': ['a', 'b'],
                }
            ),
            content_type='application/json',
        )
        assert response.status_code == 200
        assert len(response.get_data()) > 0
        assert list(tmp_path.iterdir()) == []


class TestPreviewTruncationDoesNotAffectExports:
    """P2.2 / P5 - the preview is capped; csv_data and exports are not."""

    LONG = 'L' * 2000

    def _process(self, client):
        payload = json.dumps(
            [{'text': self.LONG, 'items': list(range(100)), 'meta': {'a': self.LONG}}]
        )
        return json.loads(
            client.post(
                '/process',
                data={
                    'input_method': 'paste',
                    'pasted_json': payload,
                    'json_path': '(root)',
                },
            ).data
        )

    def test_preview_is_truncated(self, client):
        preview = self._process(client)['preview'][0]
        assert preview['text'].endswith('… (truncated)')
        assert len(preview['text']) < len(self.LONG)
        assert len(preview['items']) == 21

    def test_csv_data_keeps_full_fidelity(self, client):
        csv_data = self._process(client)['csv_data'][0]
        assert csv_data['text'] == self.LONG
        assert csv_data['meta.a'] == self.LONG
        assert json.loads(csv_data['items']) == list(range(100))

    def test_server_csv_export_is_untruncated(self, client):
        data = self._process(client)
        response = client.post(
            '/export-csv',
            data=json.dumps({'csv_data': data['csv_data'], 'csv_columns': data['csv_columns']}),
            content_type='application/json',
        )
        rows = list(csv.reader(io.StringIO(response.get_data(as_text=True))))
        assert self.LONG in rows[1]
        assert '… (truncated)' not in response.get_data(as_text=True)

    def test_xlsx_export_is_untruncated(self, client):
        from openpyxl import load_workbook

        data = self._process(client)
        response = client.post(
            '/export-xlsx',
            data=json.dumps({'csv_data': data['csv_data'], 'csv_columns': data['csv_columns']}),
            content_type='application/json',
        )
        ws = load_workbook(io.BytesIO(response.data)).active
        values = [cell.value for cell in ws[2]]
        assert self.LONG in values


class TestApiFetchDecodesWithoutAnExtraCopy:
    """P12 - the streamed bytearray is decoded directly."""

    @patch('routes.requests.get')
    @patch('routes.validate_url')
    def test_bytearray_is_decoded_in_place(self, mock_validate, mock_get, client):
        mock_validate.return_value = (True, None)
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        # Several chunks, plus a multi-byte character split across none of them,
        # to prove the accumulated bytearray is what gets decoded.
        mock_resp.iter_content.return_value = [
            b'[{"name": "Zo',
            'ë'.encode(),
            b'"}]',
        ]
        mock_resp.raise_for_status.return_value = None
        mock_get.return_value = mock_resp

        response = client.post(
            '/process',
            data={
                'input_method': 'api',
                'api_url': 'https://api.example.com/data',
                'json_path': '(root)',
            },
        )
        data = json.loads(response.data)
        assert data['success'] is True
        assert data['csv_data'][0]['name'] == 'Zoë'


class TestRateLimitTopologyGuard:
    """
    2.10 / F12 - memory:// counters are process-local, so the guard must fail
    closed rather than let a multi-worker deployment silently enforce N x the
    configured limit.
    """

    def test_memory_storage_counters_are_not_shared(self):
        """
        The multiplier is real, not theoretical.

        Two limiter storages on memory:// do not see each other's hits, which is
        exactly what happens across gunicorn workers and across replicas.
        """
        from limits.storage import storage_from_string

        first = storage_from_string('memory://')
        second = storage_from_string('memory://')

        for _ in range(5):
            first.incr('shared-key', 60)

        assert first.get('shared-key') == 5
        assert second.get('shared-key') == 0

    def test_storage_uri_is_configurable_at_all(self, fresh_config):
        # config.py hardcoded 'memory://' before v1.2, so no deployment could set
        # shared storage even if it wanted to (2.10a).
        cfg = fresh_config(RATELIMIT_STORAGE_URI='redis://localhost:6379/0')
        assert cfg.RATELIMIT_STORAGE_URI == 'redis://localhost:6379/0'

    def test_default_topology_starts_clean(self, fresh_config, caplog):
        cfg = fresh_config(APP_ENV=None, WEB_CONCURRENCY=None, APP_REPLICAS=None)
        with caplog.at_level(logging.WARNING):
            create_app(cfg)
        assert 'Rate-limit topology' not in caplog.text

    def test_multi_worker_on_memory_storage_warns_outside_production(self, fresh_config, caplog):
        cfg = fresh_config(APP_ENV=None, WEB_CONCURRENCY='4', APP_REPLICAS='1')
        with caplog.at_level(logging.WARNING):
            create_app(cfg)
        assert 'process-local' in caplog.text
        assert '4 x 1' in caplog.text

    def test_multi_worker_on_memory_storage_raises_in_production(self, fresh_config):
        cfg = fresh_config(
            APP_ENV='production',
            SECRET_KEY='k',
            WEB_CONCURRENCY='4',
            APP_REPLICAS='1',
        )
        with pytest.raises(RuntimeError, match='process-local'):
            create_app(cfg)

    def test_multi_replica_on_memory_storage_raises_in_production(self, fresh_config):
        cfg = fresh_config(
            APP_ENV='production',
            SECRET_KEY='k',
            WEB_CONCURRENCY='1',
            APP_REPLICAS='3',
        )
        with pytest.raises(RuntimeError, match='1 x 3'):
            create_app(cfg)

    def test_missing_declaration_raises_in_production(self, fresh_config):
        cfg = fresh_config(APP_ENV='production', SECRET_KEY='k', WEB_CONCURRENCY=None)
        with pytest.raises(RuntimeError, match='WEB_CONCURRENCY is not declared'):
            create_app(cfg)

    def test_missing_replica_declaration_raises_in_production(self, fresh_config):
        cfg = fresh_config(
            APP_ENV='production', SECRET_KEY='k', WEB_CONCURRENCY='1', APP_REPLICAS=None
        )
        with pytest.raises(RuntimeError, match='APP_REPLICAS is not declared'):
            create_app(cfg)

    def test_shared_storage_allows_a_declared_multi_worker_topology(self, fresh_config):
        cfg = fresh_config(
            APP_ENV='production',
            SECRET_KEY='k',
            WEB_CONCURRENCY='4',
            APP_REPLICAS='2',
            RATELIMIT_STORAGE_URI='redis://localhost:6379/0',
        )
        # Constructed, not connected: Flask-Limiter dials Redis lazily.
        app = create_app(cfg)
        assert app.config['RATELIMIT_STORAGE_URI'].startswith('redis://')

    def test_start_command_contradicting_the_declaration_raises(self, fresh_config):
        cfg = fresh_config(
            APP_ENV='production', SECRET_KEY='k', WEB_CONCURRENCY='1', APP_REPLICAS='1'
        )
        app = create_app(cfg)
        argv = ['/usr/bin/gunicorn', 'app:create_app()', '--workers', '4']
        with pytest.raises(RuntimeError, match='start command runs --workers 4'):
            check_rate_limit_topology(app, argv=argv)

    def test_start_command_agreeing_with_the_declaration_is_fine(self, fresh_config):
        cfg = fresh_config(
            APP_ENV='production',
            SECRET_KEY='k',
            WEB_CONCURRENCY='4',
            APP_REPLICAS='1',
            RATELIMIT_STORAGE_URI='redis://localhost:6379/0',
        )
        app = create_app(cfg)
        check_rate_limit_topology(app, argv=['/usr/bin/gunicorn', '--workers=4'])

    def test_non_gunicorn_argv_is_ignored(self, fresh_config):
        cfg = fresh_config()
        app = create_app(cfg)
        # pytest's own -w-like flags must not be read as a worker declaration.
        check_rate_limit_topology(app, argv=['pytest', '-w', '9'])

    def test_worker_count_parsing(self):
        assert worker_count_from_start_command(['gunicorn', '--workers', '3']) == 3
        assert worker_count_from_start_command(['gunicorn', '-w', '2']) == 2
        assert worker_count_from_start_command(['/x/gunicorn', '--workers=7']) == 7
        assert worker_count_from_start_command(['gunicorn', '--bind', ':80']) is None
        assert worker_count_from_start_command(['gunicorn', '--workers', 'x']) is None
        assert worker_count_from_start_command([]) is None
