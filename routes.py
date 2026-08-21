"""Flask route handlers."""

import csv
import io
import json
import logging
import re

import requests
from flask import (
    Blueprint,
    Response,
    current_app,
    jsonify,
    render_template,
    request,
    send_file,
)
from requests.auth import HTTPBasicAuth
from werkzeug.exceptions import HTTPException

from config import DEFAULT_MAX_EXPORT_CELLS
from extensions import limiter
from helpers import (
    extract_by_path,
    extract_table_data,
    flatten_rows,
    get_all_columns,
    is_formula_trigger,
    parse_jsonl,
    preview_truncate,
    sanitize_cell,
    serialize_cell_value,
)
from security import validate_url

logger = logging.getLogger(__name__)

bp = Blueprint('main', __name__)


# F4: outbound header names come from the client, so this is a real allowlist,
# not a token regex plus a list of names to reject. A denylist cannot work here:
# HTTP field names are case-insensitive, so `Host`, `PROXY-AUTHORIZATION` and
# `CoNnEcTiOn` all slip past a lowercase membership test, and anything simply
# absent from the list would pass. Names are stripped and lowercased before the
# membership test; the regex stays only as a syntax check on top of it.
ALLOWED_OUTBOUND_HEADERS = frozenset(
    {
        'accept',
        'accept-language',
        'authorization',
        'user-agent',
        'x-api-key',
    }
)

HEADER_NAME_PATTERN = re.compile(r'^[A-Za-z0-9-]+$')

# F13: `accept=".json,.jsonl"` on the file input is client-side only. The
# extension check is the authoritative one; the content-type check is deliberately
# lenient because browsers send application/octet-stream (or nothing at all) for
# extensions they do not recognize -- .jsonl in particular -- so a strict list
# would reject legitimate uploads.
ALLOWED_UPLOAD_EXTENSIONS = ('.json', '.jsonl')

# Rows buffered before a chunk of CSV is handed to the WSGI server.
CSV_STREAM_CHUNK_ROWS = 500

ALLOWED_UPLOAD_CONTENT_TYPES = frozenset(
    {
        '',
        'application/json',
        'application/jsonl',
        'application/ld+json',
        'application/octet-stream',
        'application/x-ndjson',
        'text/json',
        'text/plain',
        'text/x-json',
    }
)


def validate_upload(file_storage):
    """Return an error message for a file we will not try to parse, else None."""
    filename = (file_storage.filename or '').strip().lower()
    if not filename.endswith(ALLOWED_UPLOAD_EXTENSIONS):
        return 'File must be a .json or .jsonl file'

    content_type = (file_storage.mimetype or '').strip().lower()
    if content_type not in ALLOWED_UPLOAD_CONTENT_TYPES:
        return f'Unsupported content type: {content_type}'

    return None


def is_allowed_outbound_header(name):
    """True when a client-supplied outbound header name may be forwarded."""
    normalized = name.strip().lower()
    if not HEADER_NAME_PATTERN.match(normalized):
        return False
    return normalized in ALLOWED_OUTBOUND_HEADERS


@bp.route('/')
def index():
    """Render the main page."""
    return render_template('index.html')


@bp.route('/health')
def health():
    """Health check endpoint."""
    payload = {'status': 'ok'}
    if current_app.config.get('HEALTH_REVEAL_VERSION', True):
        payload['version'] = current_app.config['APP_VERSION']
    return jsonify(payload)


@bp.route('/process', methods=['POST'])
@limiter.limit(lambda: current_app.config.get('RATE_LIMIT_PROCESS', '30/minute'))
def process_json():
    """Process JSON data from file upload, pasted text, or API fetch."""
    try:
        input_method = request.form.get('input_method')
        json_data = None

        data_format = request.form.get('data_format', 'json')

        if input_method == 'file':
            if 'json_file' not in request.files:
                return jsonify({'error': 'No file uploaded'}), 400
            file = request.files['json_file']
            if file.filename == '':
                return jsonify({'error': 'No file selected'}), 400
            upload_error = validate_upload(file)
            if upload_error:
                return jsonify({'error': upload_error}), 400
            try:
                content = file.read().decode('utf-8')
                json_data = parse_jsonl(content) if data_format == 'jsonl' else json.loads(content)
            except UnicodeDecodeError:
                return jsonify({'error': 'File must be UTF-8 encoded'}), 400
            except (json.JSONDecodeError, ValueError) as e:
                return jsonify({'error': f'Invalid data in file: {str(e)}'}), 400

        elif input_method == 'paste':
            pasted_json = request.form.get('pasted_json', '').strip()
            if not pasted_json:
                return jsonify({'error': 'No JSON provided'}), 400
            try:
                if data_format == 'jsonl':
                    json_data = parse_jsonl(pasted_json)
                else:
                    json_data = json.loads(pasted_json)
            except (json.JSONDecodeError, ValueError) as e:
                return jsonify({'error': f'Invalid data: {str(e)}'}), 400

        elif input_method == 'api':
            api_url = request.form.get('api_url', '').strip()
            if not api_url:
                return jsonify({'error': 'No API URL provided'}), 400

            is_valid, error_msg = validate_url(api_url)
            if not is_valid:
                return jsonify({'error': error_msg}), 400

            auth_method = request.form.get('auth_method', 'none')
            headers = {}
            auth = None
            params = {}

            if auth_method == 'api_key':
                header_name = request.form.get('api_key_header', 'X-API-Key')
                api_key = request.form.get('api_key', '')
                if api_key:
                    if not is_allowed_outbound_header(header_name):
                        return jsonify(
                            {
                                'error': 'Header name is not permitted. Allowed: '
                                + ', '.join(sorted(ALLOWED_OUTBOUND_HEADERS))
                            }
                        ), 400
                    headers[header_name.strip()] = api_key
            elif auth_method == 'basic':
                username = request.form.get('basic_username', '')
                password = request.form.get('basic_password', '')
                if username:
                    auth = HTTPBasicAuth(username, password)
            elif auth_method == 'bearer':
                bearer_token = request.form.get('bearer_token', '')
                if bearer_token:
                    headers['Authorization'] = f'Bearer {bearer_token}'
            elif auth_method == 'query_param':
                param_name = request.form.get('query_param_name', 'api_key')
                param_value = request.form.get('query_param_value', '')
                if param_value:
                    params[param_name] = param_value

            try:
                timeout = current_app.config['API_FETCH_TIMEOUT']
                max_size = current_app.config['API_FETCH_MAX_RESPONSE']

                # Use original URL to preserve TLS/SNI verification.
                # SSRF mitigated by: pre-request DNS validation + disabled redirects.
                # Residual DNS rebinding risk is minimal (requires attacker-controlled
                # DNS with sub-millisecond TTL between our check and requests' connect).
                resp = requests.get(
                    api_url,
                    headers=headers,
                    auth=auth,
                    params=params,
                    timeout=timeout,
                    stream=True,
                    allow_redirects=False,
                )
                resp.raise_for_status()

                content = bytearray()
                for chunk in resp.iter_content(chunk_size=8192):
                    content.extend(chunk)
                    if len(content) > max_size:
                        return jsonify(
                            {
                                'error': f'API response exceeds maximum size '
                                f'({max_size // (1024 * 1024)}MB)'
                            }
                        ), 400

                # bytearray decodes directly; bytes(content) made a second full
                # copy of the response body at peak (P12). Parsing, flattening and
                # jsonify still materialize the dataset -- this removes one copy,
                # it does not make the pipeline low-memory.
                text = content.decode('utf-8')
                json_data = parse_jsonl(text) if data_format == 'jsonl' else json.loads(text)

            except requests.exceptions.Timeout:
                return jsonify({'error': 'API request timed out'}), 400
            except requests.exceptions.RequestException:
                # Fixed message, no interpolation: requests' exception text carries
                # the full URL, and the query string, fragment, userinfo AND path
                # can each hold a token (F3/F9). Redacting one component is not
                # enough, so nothing user-controlled is logged at all.
                logger.warning('API request failed')
                return jsonify({'error': 'API request failed'}), 400
            except json.JSONDecodeError:
                return jsonify({'error': 'API response is not valid JSON'}), 400
            except ValueError:
                # parse_jsonl raises ValueError on a malformed line. Without this
                # it reached the outer handler as a 500 with a logged traceback,
                # although it is the caller's data that is wrong (F9).
                return jsonify({'error': 'API response is not valid JSONL'}), 400
        else:
            return jsonify({'error': 'Invalid input method'}), 400

        # Check if user selected a specific JSON path
        json_path = request.form.get('json_path', '')

        if json_path:
            selected = extract_by_path(json_data, json_path)
            if selected is None:
                return jsonify({'error': f'Path "{json_path}" not found'}), 400
            if isinstance(selected, list):
                table_data = extract_table_data(
                    selected, max_depth=current_app.config['FLATTEN_MAX_DEPTH']
                )
            elif isinstance(selected, dict):
                table_data = [selected]
            else:
                return jsonify(
                    {'error': f'Path "{json_path}" is a primitive value; pick an object or array'}
                ), 400
        else:
            # No path chosen yet — let the client render a tree picker
            return jsonify({'needs_selection': True, 'raw_json': json_data})

        if not table_data:
            return jsonify({'error': 'Could not extract tabular data from JSON'}), 400

        columns = get_all_columns(table_data)
        preview_limit = current_app.config['PREVIEW_ROW_LIMIT']
        # A separate projection, not a mutation: csv_data below is built from the
        # untouched rows, so exports stay full-fidelity (P2.2/P5).
        preview_data = [preview_truncate(row) for row in table_data[:preview_limit]]

        max_depth = current_app.config['FLATTEN_MAX_DEPTH']
        # One pass instead of flatten-then-rescan: names are collected into a set
        # while flattening and sorted once at the end, which is byte-identical to
        # get_all_columns' sorted output (P8).
        csv_data, csv_columns = flatten_rows(table_data, max_depth=max_depth)

        # Additive only: no existing key changes name, type or meaning.
        # total_cells/max_export_cells let the client grey out the Excel entry
        # BEFORE the user clicks, rather than after a 400 (D6).
        return jsonify(
            {
                'success': True,
                'columns': columns,
                'preview': preview_data,
                'total_rows': len(table_data),
                'total_cells': len(csv_data) * len(csv_columns),
                'max_export_cells': current_app.config.get(
                    'MAX_EXPORT_CELLS', DEFAULT_MAX_EXPORT_CELLS
                ),
                'csv_data': csv_data,
                'csv_columns': csv_columns,
            }
        )

    except HTTPException:
        # Werkzeug raises these lazily inside the route -- RequestEntityTooLarge
        # fires the first time the oversized body is read. They already carry the
        # right status, so let Flask's error handlers render them as JSON (F10)
        # instead of swallowing them into a 500 below.
        raise
    except RecursionError:
        # Valid JSON can nest deeply enough to exhaust the C stack, in json.loads
        # itself, in the helpers, or in the response encoder. That is the caller's
        # document, so it is a 400 -- and it is not worth an exception traceback
        # in the logs (F8).
        return jsonify({'error': 'JSON nesting too deep'}), 400
    except Exception:
        logger.exception('Unexpected error in process_json')
        return jsonify({'error': 'An internal error occurred'}), 500


def _stream_csv(columns, rows):
    """
    Yield the CSV a chunk of rows at a time.

    csv.writer needs a text buffer, so one StringIO is reused and truncated every
    CSV_STREAM_CHUNK_ROWS rows instead of the whole file being built in memory
    before the first byte goes out (P3).
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    def drain():
        value = buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)
        return value

    writer.writerow([sanitize_cell(column) for column in columns])
    yield drain()

    for index, row in enumerate(rows, start=1):
        writer.writerow([sanitize_cell(row.get(column, '')) for column in columns])
        if index % CSV_STREAM_CHUNK_ROWS == 0:
            yield drain()

    remainder = drain()
    if remainder:
        yield remainder


def _append_xlsx_row(ws, values):
    """
    Append one row to a worksheet, writing formula-triggering strings as strings.

    openpyxl serializes a str starting with '=' as a formula cell, so Excel would
    evaluate an attacker-supplied value on open (F1). XLSX carries an explicit
    type per cell, so the fix is to pin data_type rather than mangle the text the
    way the delimited exports have to.
    """
    serialized = [serialize_cell_value(value) for value in values]
    ws.append(serialized)
    row_index = ws.max_row
    for column_index, value in enumerate(serialized, start=1):
        if is_formula_trigger(value):
            ws.cell(row=row_index, column=column_index).data_type = 's'


@bp.route('/export-csv', methods=['POST'])
@limiter.limit(lambda: current_app.config.get('RATE_LIMIT_EXPORT', '60/minute'))
def export_csv():
    """Export data as CSV file."""
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({'error': 'Invalid or missing JSON body'}), 400

        csv_data = data.get('csv_data', [])
        csv_columns = data.get('csv_columns', [])

        if not csv_data:
            return jsonify({'error': 'No data to export'}), 400

        # Streamed, and deliberately uncapped: CSV is natively streamable with no
        # temp files, so every dataset /process accepts stays exportable by this
        # route even when it is too large for a workbook (P3/D6). That, not an
        # unbounded XLSX path, is what keeps the export contract as wide as the
        # input contract.
        return Response(
            _stream_csv(csv_columns, csv_data),
            mimetype='text/csv',
            headers={
                'Content-Disposition': 'attachment; filename=exported_data.csv',
                'Content-Type': 'text/csv; charset=utf-8',
            },
        )

    except HTTPException:
        raise
    except Exception:
        logger.exception('Unexpected error in export_csv')
        return jsonify({'error': 'Export failed'}), 500


@bp.route('/export-xlsx', methods=['POST'])
@limiter.limit(lambda: current_app.config.get('RATE_LIMIT_EXPORT', '60/minute'))
def export_xlsx():
    """Export data as Excel file."""
    try:
        from openpyxl import Workbook

        data = request.get_json(silent=True)
        if not data:
            return jsonify({'error': 'Invalid or missing JSON body'}), 400

        xlsx_data = data.get('csv_data', [])
        xlsx_columns = data.get('csv_columns', [])

        if not xlsx_data:
            return jsonify({'error': 'No data to export'}), 400

        limit = current_app.config.get('MAX_EXPORT_CELLS', DEFAULT_MAX_EXPORT_CELLS)
        cells = len(xlsx_data) * len(xlsx_columns)
        if limit and cells > limit:
            # Defence in depth for direct API callers; the UI already greyed the
            # Excel entry out using total_cells/max_export_cells from /process.
            # A refusal, never a silent truncation -- a partial spreadsheet is
            # worse than none.
            return jsonify(
                {
                    'error': (
                        f'Dataset is {cells} cells, above the Excel export limit '
                        f'of {limit}; export CSV or TSV instead.'
                    )
                }
            ), 400

        wb = Workbook()
        ws = wb.active
        ws.title = 'Data'

        _append_xlsx_row(ws, xlsx_columns)
        for row in xlsx_data:
            _append_xlsx_row(ws, [row.get(col, '') for col in xlsx_columns])

        # Normal-mode Workbook and a plain BytesIO: no OS temp files anywhere.
        # openpyxl's write_only mode writes worksheet parts to disk, and
        # SpooledTemporaryFile is either pointless (its default max_size=0 never
        # rolls over, so it is a BytesIO with extra indirection) or disk-backed
        # (a non-zero threshold, or any fileno() call, puts payload bytes on
        # disk). The cell budget above is what bounds memory (D6).
        output = io.BytesIO()
        wb.save(output)
        del wb
        output.seek(0)

        # send_file streams the buffer out in chunks; getvalue() would make a
        # second full copy of the workbook at peak.
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='exported_data.xlsx',
        )

    except HTTPException:
        raise
    except Exception:
        logger.exception('Unexpected error in export_xlsx')
        return jsonify({'error': 'Export failed'}), 500
